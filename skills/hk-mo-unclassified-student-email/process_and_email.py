"""
港澳地区未分班超14天明细同步教务
Step 1: 处理 Excel - 筛选未分班学员
Step 2: 登录腾讯企业邮箱 - 发送邮件保存草稿
"""
import os
import sys
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
AUTH_STATE_FILE = SCRIPT_DIR / "email_auth_state.json"

SOURCE_SHEET = "海外未分班学员明细"

FILTER_COL_STAGE = "学员阶段"          # B列
FILTER_VAL_STAGE = "组班中"

FILTER_COL_CHANNEL = "渠道一级分类"     # G列
FILTER_VAL_CHANNEL = "海外港澳商务"

FILTER_COL_WAIT = "未退费等待时长"      # R列
FILTER_WAIT_THRESHOLD = 14

RECIPIENTS_TO = ["recipient1@example.com", "recipient2@example.com"]
RECIPIENTS_CC = ["cc1@example.com", "cc2@example.com", "cc3@example.com"]


def make_output_name():
    today = datetime.now().strftime("%Y.%m.%d")
    return f"【港澳海外商务】等班14天以上学员明细+{today}"


def parse_wait_days(value):
    if pd.isna(value):
        return 0
    text = str(value)
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else 0


def step1_process_excel(source_path: str) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_name = make_output_name()

    wb = load_workbook(source_path)
    if SOURCE_SHEET not in wb.sheetnames:
        print(f"[错误] 找不到 sheet: {SOURCE_SHEET}")
        print(f"  可用 sheets: {wb.sheetnames}")
        sys.exit(1)

    df = pd.read_excel(source_path, sheet_name=SOURCE_SHEET, header=4)
    print(f"[原始数据] {len(df)} 行")

    mask = df[FILTER_COL_STAGE] == FILTER_VAL_STAGE
    print(f"  学员阶段=组班中: {mask.sum()} 行")

    mask &= df[FILTER_COL_CHANNEL] == FILTER_VAL_CHANNEL
    print(f"  + 渠道=海外港澳商务: {mask.sum()} 行")

    days = df[FILTER_COL_WAIT].apply(parse_wait_days)
    mask &= days > FILTER_WAIT_THRESHOLD
    print(f"  + 等待时长>14天: {mask.sum()} 行")

    filtered = df[mask].reset_index(drop=True)
    if filtered.empty:
        print("[警告] 筛选结果为空，无符合条件的学员")

    output_path = OUTPUT_DIR / f"{output_name}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        filtered.to_excel(writer, sheet_name=output_name[:31], index=False)
    print(f"[完成] 输出文件: {output_path}")
    print(f"  共 {len(filtered)} 条记录")
    return output_path


def df_to_html_table(df: pd.DataFrame) -> str:
    return df.to_html(index=False, border=1, na_rep="")


# ─── STEP 2: 邮件发送 ───

def step2_send_email(excel_path: Path):
    from playwright.sync_api import sync_playwright
    import time

    today = datetime.now().strftime("%Y.%m.%d")
    subject = f"【港澳海外商务】KOL加急排课及等班14天以上学员明细-{today}"

    df = pd.read_excel(excel_path)
    table_html = df_to_html_table(df)

    body_html = (
        "<p>Dear教务老师：</p>"
        "<p>&nbsp;&nbsp;&nbsp;&nbsp;海外港澳商务渠道-等班≥14天的学员明细，"
        "辛苦帮忙优先排课，谢谢~</p><br>"
    ) + table_html

    with sync_playwright() as p:
        if AUTH_STATE_FILE.exists():
            context = p.chromium.launch(headless=False).new_context(
                storage_state=str(AUTH_STATE_FILE)
            )
            print("[登录] 使用已保存的 cookie")
        else:
            context = p.chromium.launch(headless=False).new_context()
            print("[登录] 首次登录，请手动输入账号密码")

        page = context.new_page()
        page.goto("https://exmail.qq.com/login", timeout=60000)

        if not AUTH_STATE_FILE.exists():
            print("[提示] 请在浏览器中完成登录...")
            page.wait_for_url("**/cgi-bin/frame_html*", timeout=120000)
            context.storage_state(path=str(AUTH_STATE_FILE))
            print("[登录] cookie 已保存")
        else:
            time.sleep(5)
            if "login" in page.url:
                print("[提示] cookie 已过期，请重新登录...")
                page.wait_for_url("**/cgi-bin/frame_html*", timeout=120000)
                context.storage_state(path=str(AUTH_STATE_FILE))
                print("[登录] cookie 已更新")

        print("[邮箱] 已进入主页面，准备写信...")

        page.click('a:has-text("写信")')
        mf = page.frame("mainFrame")
        mf.wait_for_url("**/readtemplate*t=compose*", timeout=15000)
        time.sleep(2)
        print("[写信] 写信页面已加载")

        for addr in RECIPIENTS_TO:
            to_input = mf.locator("#toAreaCtrl .addr_text input")
            to_input.click()
            to_input.fill(addr)
            to_input.press("Enter")
            time.sleep(0.5)

        mf.locator("#aCC").click()
        time.sleep(0.5)
        for addr in RECIPIENTS_CC:
            cc_input = mf.locator("#ccAreaCtrl .addr_text input")
            cc_input.click()
            cc_input.fill(addr)
            cc_input.press("Enter")
            time.sleep(0.5)

        subj_input = mf.locator('input#subject[name="subject"]')
        subj_input.click()
        subj_input.fill(subject)
        print(f"[主题] {subject}")

        editor_frame = mf.frame_locator("#QMEditorIfrmEditArea")
        editor_body = editor_frame.locator("body")
        editor_body.evaluate("(el, html) => { el.innerHTML = html; }", body_html)
        print("[正文] 已写入")

        file_input = mf.locator('input[name="UploadFile"]').first
        file_input.set_input_files(str(excel_path))
        time.sleep(3)
        print(f"[附件] 已上传: {excel_path.name}")

        mf.locator('input[name="savebtn"]').first.click()
        time.sleep(2)
        print("[完成] 邮件已保存到草稿箱")

        context.storage_state(path=str(AUTH_STATE_FILE))
        context.close()


# ─── MAIN ───

def main():
    default_path = SCRIPT_DIR / "sample" / "海外未分班学员明细.xlsx"
    source_path = sys.argv[1] if len(sys.argv) > 1 else str(default_path)

    if not os.path.exists(source_path):
        print(f"[错误] 文件不存在: {source_path}")
        sys.exit(1)

    output_path = step1_process_excel(source_path)
    step2_send_email(output_path)


if __name__ == "__main__":
    main()
