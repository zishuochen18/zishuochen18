"""
TMK 周报自动生成脚本

使用方法：
   1. 将本周下载的 3 个 Excel 文件放到 zhoubao/ 下的某个子文件夹中
   2. 运行：python generate_weekly_report.py zhoubao/5.22测试
   3. 生成的 Markdown 周报在同一文件夹下

数据源：
   - 海外TMK做工监控.xlsx       （主表，含通次/通时/跟进时效/新老生跟进/结果指标）
   - 海外TMK未邀约做工监控播报.xlsx   （含进线非勿扰时段相关）
   - 海外TMK做工勿扰情况汇总.xlsx     （勿扰跟进、约课率细节）

输出：
   - 周报.md（含分组表头数据表 + 要点式文字分析）
"""

import os
import sys
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ALERT_THRESHOLD = -0.10  # 环比下降超 10% 视为异常

# 标准表格列定义：(分组名, 列名, 来源文件关键词, 数据格式)
# 数据格式：num=数字, pct=百分比, int=整数
STANDARD_COLUMNS = [
    ("TMK",       "日均通次",                   "做工监控", "int"),
    ("TMK",       "日均通次环比",               "做工监控", "pct"),
    ("TMK个人做工", "日均通时（分）",             "做工监控", "num"),
    ("TMK个人做工", "日均通时环比",               "做工监控", "pct"),
    ("TMK个人做工", "昨日通次",                   "做工监控", "int"),
    ("TMK个人做工", "昨日通时",                   "做工监控", "int"),
    ("跟进时效",   "昨日生均跟进时效（分）",       "做工监控", "num"),
    ("跟进时效",   "生均跟进时效（分）",           "做工监控", "num"),
    ("跟进时效",   "生均跟进时效达成率",           "做工监控", "pct"),
    ("新生跟进",   "拨打学员数",                  "做工监控", "int"),
    ("新生跟进",   "生均通次",                    "做工监控", "num"),
    ("新生跟进",   "生均通时（分）",              "做工监控", "num"),
    ("新生跟进",   "有效接通率",                  "做工监控", "pct"),
    ("老生跟进",   "同期拨打环比",                "做工监控", "pct"),
    ("老生跟进",   "生均通次",                    "做工监控", "num"),
    ("老生跟进",   "批量拨打次数",                "做工监控", "int"),
    ("老生跟进",   "有效接通率",                  "做工监控", "pct"),
    ("结果指标",   "思维例子数",                  "做工监控", "int"),
    ("结果指标",   "思维约课数",                  "做工监控", "int"),
    ("结果指标",   "首发本组约课数",              "做工监控", "int"),
    ("结果指标",   "思维例子约课率",              "做工监控", "pct"),
    ("新生跟进(进线非勿扰)", "生均通次",          "未邀约",   "num"),
    ("新生跟进(进线非勿扰)", "生均通次（进线非勿扰时段的例子）", "未邀约", "num"),
    ("新生跟进(进线非勿扰)", "生均通时（分）",     "未邀约",   "num"),
    ("新生跟进(进线非勿扰)", "首次接通邀约率",     "勿扰",     "pct"),
]


def find_excel(folder: str, keyword: str) -> str:
    for f in os.listdir(folder):
        if keyword in f and f.endswith(".xlsx"):
            return os.path.join(folder, f)
    return ""


def parse_filter_params(df: pd.DataFrame, max_rows: int = 4) -> dict:
    """从 Excel 前几行提取筛选参数"""
    params = {}
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i].tolist()
        for j, val in enumerate(row):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str in ("开始日期", "开始时间："):
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params["start_date"] = str(row[j + 1]).split(" ")[0]
                elif val_str in ("结束日期", "结束时间："):
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params["end_date"] = str(row[j + 1]).split(" ")[0]
                elif val_str == "一级渠道：":
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params["channel"] = str(row[j + 1])
                elif val_str == "二级渠道：":
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params["sub_channel"] = str(row[j + 1])
                elif val_str == "一级渠道":
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params.setdefault("channel", str(row[j + 1]))
                elif val_str == "二级渠道":
                    if j + 1 < len(row) and pd.notna(row[j + 1]):
                        params.setdefault("sub_channel", str(row[j + 1]))
    return params


def parse_excel_with_groups(filepath: str, header_group_row: int, header_name_row: int,
                             data_start_row: int) -> tuple:
    """
    通用解析：根据分组表头行和列名行解析 Excel
    返回 (params, df, column_meta)
    column_meta = [(col_index, group, name), ...]
    """
    raw = pd.read_excel(filepath, sheet_name="Sheet1", header=None)
    params = parse_filter_params(raw)

    group_row = raw.iloc[header_group_row].tolist()
    name_row = raw.iloc[header_name_row].tolist()

    column_meta = []
    current_group = ""
    for i in range(len(name_row)):
        if pd.notna(group_row[i]):
            current_group = str(group_row[i]).strip()
        if pd.notna(name_row[i]):
            column_meta.append((i, current_group, str(name_row[i]).strip()))

    # metric_start = 列名行中第一个非空列的位置（最可靠的边界）
    metric_start = column_meta[0][0] if column_meta else 2

    # 构建 dataframe
    columns = ["_skip"] * len(raw.columns)
    if metric_start >= 2:
        columns[metric_start - 2] = "TMK小组"
    columns[metric_start - 1] = "TMK"
    for idx, group, name in column_meta:
        if idx >= metric_start:
            full_name = f"{group}__{name}"
            columns[idx] = full_name

    # 去重
    seen = {}
    unique_columns = []
    for c in columns:
        if c in seen:
            seen[c] += 1
            unique_columns.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            unique_columns.append(c)

    data = raw.iloc[data_start_row:].copy()
    data.columns = unique_columns[:len(data.columns)]
    data = data.reset_index(drop=True)

    # 转数值列
    for col in data.columns:
        if col != "_skip" and col != "TMK小组" and col != "TMK":
            try:
                data[col] = pd.to_numeric(data[col], errors="coerce")
            except Exception:
                pass

    # 过滤底部口径说明等无效行
    if "TMK" in data.columns:
        data = data[
            data["TMK"].notna()
            & ~data["TMK"].astype(str).str.match(r'^\d+）')
            & ~data["TMK"].astype(str).str.contains("口径|说明", na=False)
        ].reset_index(drop=True)

    # TMK小组列向下填充（Excel 中组名只在每组第一行）
    if "TMK小组" in data.columns:
        data["TMK小组"] = data["TMK小组"].ffill()

    return params, data, column_meta


def parse_work_monitor(filepath: str) -> tuple:
    """海外TMK做工监控：行4分组、行5列名、行6数据"""
    return parse_excel_with_groups(filepath, 4, 5, 6)


def parse_uninvited(filepath: str) -> tuple:
    """海外TMK未邀约做工监控播报：行6分组、行7列名、行8数据"""
    return parse_excel_with_groups(filepath, 6, 7, 8)


def parse_disturb(filepath: str) -> tuple:
    """海外TMK做工勿扰情况汇总：行3分组、行4列名、行5数据"""
    return parse_excel_with_groups(filepath, 3, 4, 5)


def find_column_value(row, df, group: str, name: str):
    """在指定数据行中根据 group + name 找列值"""
    target = f"{group}__{name}"
    for col in df.columns:
        if col == target or col.startswith(target + "_"):
            return row.get(col)
    return None


def fmt(val, kind: str) -> str:
    """格式化单元格值"""
    if pd.isna(val) or val is None:
        return "-"
    try:
        v = float(val)
    except (ValueError, TypeError):
        return str(val)
    if kind == "pct":
        return f"{v * 100:.2f}%"
    elif kind == "int":
        return f"{int(round(v))}"
    else:
        return f"{v:.2f}"


def find_cross_file_row(target_df, group_name: str, person_name: str):
    """
    在目标 df 中找到对应行，处理跨文件命名差异。
    - 做工监控里"团队汇总"在勿扰/未邀约文件里可能叫"总计"
    - 做工监控里"总计"+TMK小组="总计"是大盘总计
    """
    if target_df is None or "TMK" not in target_df.columns:
        return None

    # 候选名称
    name_candidates = [person_name]
    if person_name == "团队汇总":
        name_candidates.append("总计")

    # 1) 先尝试 group + name 精确匹配
    if "TMK小组" in target_df.columns and group_name:
        for n in name_candidates:
            row = target_df[
                (target_df["TMK"].astype(str).str.strip() == n)
                & (target_df["TMK小组"].astype(str).str.strip() == group_name)
            ]
            if len(row) > 0:
                return row.iloc[0]

    # 2) 仅按 name 匹配（适合个人）
    for n in name_candidates:
        row = target_df[target_df["TMK"].astype(str).str.strip() == n]
        if len(row) == 1:
            return row.iloc[0]

    return None


def merge_person_data(work_df, uninvited_df, disturb_df) -> pd.DataFrame:
    """
    合并三个数据源到一张表，包含：大盘总计 + 每组团队汇总 + 个人。
    """
    # 保留所有非空 TMK 行（含总计 / 团队汇总 / 个人）
    persons = work_df[work_df["TMK"].notna()].copy()

    rows = []
    for _, row in persons.iterrows():
        name = str(row["TMK"]).strip()
        group_name = str(row.get("TMK小组", "")).strip() if "TMK小组" in row else ""
        out = {"TMK小组": group_name, "TMK": name}

        for group, col_name, source, _ in STANDARD_COLUMNS:
            full = f"{group}__{col_name}"
            value = None

            if source == "做工监控":
                value = find_column_value(row, work_df, infer_source_group(group, col_name, "做工监控"), col_name)
            elif source == "未邀约":
                u_row = find_cross_file_row(uninvited_df, group_name, name)
                if u_row is not None:
                    src_group = infer_source_group(group, col_name, "未邀约")
                    value = find_column_value(u_row, uninvited_df, src_group, col_name)
            elif source == "勿扰":
                d_row = find_cross_file_row(disturb_df, group_name, name)
                if d_row is not None:
                    src_group = infer_source_group(group, col_name, "勿扰")
                    value = find_column_value(d_row, disturb_df, src_group, col_name)

            out[full] = value

        rows.append(out)

    return pd.DataFrame(rows)


def infer_source_group(group: str, col_name: str, source: str) -> str:
    """根据标准列的分组名，推断在源 Excel 中的分组名"""
    if source == "做工监控":
        if group == "TMK":
            return "TMK个人做工"
        return group
    elif source == "未邀约":
        if "新生跟进" in group:
            return "新生跟进"
        return group
    elif source == "勿扰":
        # 勿扰文件中：约课率/邀约率/首次接通 都在"结果指标"分组下
        # 勿扰情况分组只有：进线非勿扰数量/进线勿扰数量/勿扰占比 等
        if any(k in col_name for k in ["约课", "邀约率", "首次接通", "首次拨打", "跟进时效", "例子数"]):
            return "结果指标"
        if "勿扰" in col_name:
            return "勿扰情况"
        return "结果指标"
    return group


def export_excel_table(work_df, uninvited_df, disturb_df, output_path: str):
    """
    将个人做工数据导出为 Excel，使用合并单元格实现分组表头。
    数据按"大盘总计 → 团队汇总 → 个人"顺序展示。
    """
    merged = merge_person_data(work_df, uninvited_df, disturb_df)

    # 收集分组结构
    groups_order = []
    group_to_cols = {}
    for group, col_name, _, kind in STANDARD_COLUMNS:
        if group not in groups_order:
            groups_order.append(group)
            group_to_cols[group] = []
        group_to_cols[group].append((col_name, kind))

    wb = Workbook()
    ws = wb.active
    ws.title = "个人做工数据"

    # 样式
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    sub_header_fill = PatternFill(start_color="EEF5FF", end_color="EEF5FF", fill_type="solid")
    total_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    team_fill = PatternFill(start_color="E0E8F0", end_color="E0E8F0", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写第一行（分组表头）
    ws.cell(row=1, column=1, value="TMK小组")
    ws.cell(row=1, column=2, value="TMK")
    ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    ws.merge_cells(start_row=1, end_row=2, start_column=2, end_column=2)

    col_idx = 3
    for g in groups_order:
        n = len(group_to_cols[g])
        ws.cell(row=1, column=col_idx, value=g)
        if n > 1:
            ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + n - 1)
        col_idx += n

    # 写第二行（具体列名）
    col_idx = 3
    for g in groups_order:
        for col_name, _ in group_to_cols[g]:
            ws.cell(row=2, column=col_idx, value=col_name)
            col_idx += 1

    total_columns = col_idx - 1

    # 表头样式
    for r in [1, 2]:
        for c in range(1, total_columns + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill if r == 1 else sub_header_fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

    # 数据行（含总计 / 团队汇总 / 个人，按 merged 顺序）
    row = 3
    for _, person_row in merged.iterrows():
        ws.cell(row=row, column=1, value=person_row.get("TMK小组", "") or "")
        ws.cell(row=row, column=2, value=person_row.get("TMK", "") or "")
        col_idx = 3
        for g in groups_order:
            for col_name, kind in group_to_cols[g]:
                v = person_row.get(f"{g}__{col_name}")
                cell = ws.cell(row=row, column=col_idx)
                write_excel_cell(cell, v, kind)
                col_idx += 1

        # 高亮处理：大盘总计 / 团队汇总
        tmk_val = str(person_row.get("TMK", "")).strip()
        group_val = str(person_row.get("TMK小组", "")).strip()
        is_grand_total = tmk_val == "总计" and group_val == "总计"
        is_team_total = tmk_val == "团队汇总"

        for c in range(1, total_columns + 1):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).alignment = center
            if is_grand_total:
                ws.cell(row=row, column=c).fill = total_fill
                ws.cell(row=row, column=c).font = bold_font
            elif is_team_total:
                ws.cell(row=row, column=c).fill = team_fill
                ws.cell(row=row, column=c).font = bold_font
        row += 1

    # 自动列宽
    for c in range(1, total_columns + 1):
        col_letter = get_column_letter(c)
        max_len = 4
        for r in range(1, row):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 22)

    # 表头行高
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # 冻结前两行+前两列
    ws.freeze_panes = "C3"

    wb.save(output_path)


def write_excel_cell(cell, value, kind: str):
    """根据数据类型写入 Excel 单元格（百分比用真正的数字格式）"""
    if value is None or pd.isna(value):
        cell.value = "-"
        return
    try:
        v = float(value)
    except (ValueError, TypeError):
        cell.value = str(value)
        return

    if kind == "pct":
        cell.value = v
        cell.number_format = "0.00%"
    elif kind == "int":
        cell.value = int(round(v))
        cell.number_format = "0"
    else:
        cell.value = round(v, 2)
        cell.number_format = "0.00"


def build_html_table(work_df, uninvited_df, disturb_df) -> str:
    """构建带分组表头的 HTML 表格"""
    merged = merge_person_data(work_df, uninvited_df, disturb_df)

    # 提取总计行
    total_row_src = work_df[work_df["TMK"].astype(str).str.contains("总计", na=False)]
    total_data = {}
    if len(total_row_src) > 0:
        tr = total_row_src.iloc[0]
        for group, col_name, source, kind in STANDARD_COLUMNS:
            if source == "做工监控":
                value = find_column_value(tr, work_df, infer_source_group(group, col_name, "做工监控"), col_name)
                total_data[f"{group}__{col_name}"] = value

    # 收集分组结构（保持顺序）
    groups_order = []
    group_to_cols = {}
    for group, col_name, _, _ in STANDARD_COLUMNS:
        if group not in groups_order:
            groups_order.append(group)
            group_to_cols[group] = []
        group_to_cols[group].append(col_name)

    # 构建 HTML
    html = ['<table>', '<thead>', '<tr>']
    html.append('<th rowspan="2">TMK小组</th>')
    html.append('<th rowspan="2">TMK</th>')
    for g in groups_order:
        cols = group_to_cols[g]
        html.append(f'<th colspan="{len(cols)}">{g}</th>')
    html.append('</tr>')

    html.append('<tr>')
    for g in groups_order:
        for c in group_to_cols[g]:
            html.append(f'<th>{c}</th>')
    html.append('</tr>')
    html.append('</thead>')
    html.append('<tbody>')

    # 总计行
    html.append('<tr><td><b>总计</b></td><td><b>总计</b></td>')
    for group, col_name, _, kind in STANDARD_COLUMNS:
        v = total_data.get(f"{group}__{col_name}")
        html.append(f'<td>{fmt(v, kind)}</td>')
    html.append('</tr>')

    # 个人行
    for _, row in merged.iterrows():
        group_name = row.get("TMK小组", "") or ""
        name = row.get("TMK", "")
        html.append(f'<tr><td>{group_name}</td><td>{name}</td>')
        for group, col_name, _, kind in STANDARD_COLUMNS:
            v = row.get(f"{group}__{col_name}")
            html.append(f'<td>{fmt(v, kind)}</td>')
        html.append('</tr>')

    html.append('</tbody>')
    html.append('</table>')
    return "\n".join(html)


def generate_overall_analysis(work_df) -> list:
    """从总计行生成整体数据要点"""
    points = []
    total_rows = work_df[work_df["TMK"].astype(str).str.contains("总计", na=False)]
    if len(total_rows) == 0:
        return points
    total = total_rows.iloc[0]

    for group, col_name, source, kind in STANDARD_COLUMNS:
        if source != "做工监控":
            continue
        if col_name in ("日均通次", "日均通次环比", "日均通时（分）", "生均跟进时效达成率"):
            v = find_column_value(total, work_df, infer_source_group(group, col_name, "做工监控"), col_name)
            if pd.notna(v):
                if col_name == "日均通次":
                    points.append(f"- 团队日均通次 **{int(round(float(v)))}** 次")
                elif col_name == "日均通次环比":
                    val = float(v)
                    direction = "提升" if val >= 0 else "下降"
                    points.append(f"  - 环比{direction} {abs(val)*100:.1f}%")
                elif col_name == "日均通时（分）":
                    points.append(f"- 团队日均通时 **{float(v):.1f}** 分钟")
                elif col_name == "生均跟进时效达成率":
                    points.append(f"- 生均跟进时效达成率 **{float(v)*100:.1f}%**")

    return points


def analyze_alerts(work_df) -> list:
    """分析个人异常情况"""
    alerts = []
    persons = work_df[
        work_df["TMK"].notna()
        & ~work_df["TMK"].astype(str).str.contains("总计|团队汇总", na=False)
    ].copy()

    for _, row in persons.iterrows():
        name = str(row["TMK"]).strip()
        if not name or name == "nan":
            continue

        # 日均通次环比
        v = find_column_value(row, work_df, "TMK个人做工", "日均通次环比")
        if pd.notna(v) and float(v) < ALERT_THRESHOLD:
            alerts.append(f"- ⚠️ **{name}**：日均通次环比下降 {abs(float(v))*100:.0f}%，需关注做工量")

        # 日均通时环比
        v = find_column_value(row, work_df, "TMK个人做工", "日均通时环比")
        if pd.notna(v) and float(v) < ALERT_THRESHOLD:
            alerts.append(f"- ⚠️ **{name}**：日均通时环比下降 {abs(float(v))*100:.0f}%，通话时长缩短")

        # 跟进时效达成率
        v = find_column_value(row, work_df, "跟进时效", "生均跟进时效达成率")
        if pd.notna(v) and float(v) < 0.3:
            alerts.append(f"- ⚠️ **{name}**：跟进时效达成率仅 {float(v)*100:.0f}%，低于基准线")

    return alerts


def generate_report(folder: str) -> str:
    """主函数：生成完整周报 Markdown + Excel 数据表"""
    work_file = find_excel(folder, "做工监控")
    uninvited_file = find_excel(folder, "未邀约")
    disturb_file = find_excel(folder, "勿扰")

    if not work_file:
        print("❌ 未找到 做工监控 Excel 文件")
        return ""

    params, work_df, _ = parse_work_monitor(work_file)
    uninvited_df = None
    disturb_df = None
    if uninvited_file:
        _, uninvited_df, _ = parse_uninvited(uninvited_file)
    if disturb_file:
        _, disturb_df, _ = parse_disturb(disturb_file)

    start_date = params.get("start_date", "?")
    end_date = params.get("end_date", "?")
    channel = params.get("channel", "")
    sub_channel = params.get("sub_channel", "")

    # 导出 Excel 数据表
    excel_name = f"个人做工数据_{start_date}_{end_date}.xlsx"
    excel_path = os.path.join(folder, excel_name)
    try:
        export_excel_table(work_df, uninvited_df, disturb_df, excel_path)
        print(f">>> Excel 数据表已生成：{excel_path}")
    except PermissionError:
        print(f"❌ Excel 写入失败：{excel_path} 正被打开，请关闭后重试")
        return ""

    md = []
    md.append("# TMK 做工周报")
    md.append("")
    md.append(f"**周期**：{start_date} ~ {end_date}")
    md.append(f"**渠道**：{channel} / {sub_channel}")
    md.append(f"**生成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    md.append("")

    # 整体数据
    md.append("## 一、整体数据概览")
    md.append("")
    for p in generate_overall_analysis(work_df):
        md.append(p)
    md.append("")

    # 数据表（Excel 链接）
    md.append("## 二、个人做工数据")
    md.append("")
    md.append(f"📊 完整数据表请查看：[{excel_name}]({excel_name})")
    md.append("")

    # 个人异常
    md.append("## 三、个人异常情况")
    md.append("")
    alerts = analyze_alerts(work_df)
    if alerts:
        for a in alerts:
            md.append(a)
    else:
        md.append("- 本周无明显异常")
    md.append("")

    return "\n".join(md)


def main():
    if len(sys.argv) < 2:
        print("用法：python generate_weekly_report.py <数据文件夹路径>")
        print("示例：python generate_weekly_report.py 5.22测试")
        sys.exit(1)

    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"文件夹不存在：{folder}")
        sys.exit(1)

    sys.stdout.reconfigure(encoding="utf-8")

    print(f">>> 读取数据：{folder}")
    report = generate_report(folder)

    if report:
        output_file = os.path.join(folder, "周报.md")
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(report)
        print(f">>> 周报已生成：{output_file}")
        print()
        print(report)
    else:
        print("生成失败")


if __name__ == "__main__":
    main()

