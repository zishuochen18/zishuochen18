"""
代理对账结算自动化
Step 1: 从销售明细提取代理成交数据
Step 2: 从主订单宽表筛选本月退费订单
Step 3: 生成结算表（复制模板 + 填入数据 + 渠道汇总透视 + CPS服务费表 + 汇率）
"""
import sys
import shutil
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = Path(__file__).parent
SAMPLE_DIR = SCRIPT_DIR / "sample"
OUTPUT_DIR = SCRIPT_DIR / "output"

# 输入文件
SALES_FILE = SAMPLE_DIR / "益智海外用户销售明细_末次渠道.xlsx"
ORDERS_FILE = SAMPLE_DIR / "海外益智主订单宽表.xlsx"
TEMPLATE_FILE = SAMPLE_DIR / "香港大v Amy2026年4月渠道结算单.xlsx"

# 表头行
SALES_HEADER_ROW = 7
ORDERS_HEADER_ROW = 4

# 销售明细表列索引
SALES_COL_CHANNEL = 23  # X列：末次渠道名称
SALES_COL_CHANNEL_DATE = 6  # G列：末次渠道更新日期
SALES_COL_SIGN_TIME = 112  # 索引112：首签时间（日期时间）

# 主订单宽表列索引
ORDERS_COL_CHANNEL = 15  # P列：末次渠道名称
ORDERS_COL_REFUND_MONTH = 39  # AN列：退费月份
ORDERS_COL_REFUND_TYPE = 41  # AO列：学员退费类型
ORDERS_COL_REFUND_AMOUNT = 42  # AP列：退费金额

# 销售明细中的关键列名（在最终结算表中用）
COL_SIGN_AMOUNT = "首签金额"  # 用于渠道汇总求和
COL_TRANSIT_TIME = "进线-成交时间"  # 用于筛选 <60


def get_last_month_info():
    """返回上个月信息：(年月字符串 '2026-04', 月份数字, 显示用月份 '4月', YYYY.M.1-M.D 区间)"""
    today = datetime.now()
    first_day_this_month = today.replace(day=1)
    last_month_end = first_day_this_month - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    ym = last_month_end.strftime("%Y-%m")
    m_num = last_month_end.month
    m_display = f"{m_num}月"
    period = f"{last_month_start.year}.{last_month_start.month}.1-{last_month_end.month}.{last_month_end.day}"
    return ym, m_num, m_display, period


def get_last_working_day(year: int, month: int) -> datetime:
    """返回指定月份的最后一个工作日（遇周末向前回溯）"""
    last_day = monthrange(year, month)[1]
    d = datetime(year, month, last_day)
    while d.weekday() >= 5:  # 5=周六, 6=周日
        d -= timedelta(days=1)
    return d


def step1_extract_agent_sales(agent_name: str) -> pd.DataFrame:
    """
    Step 1: 从销售明细提取指定代理的成交数据
    1. 筛选末次渠道名称 = agent_name
    2. 计算进线-成交时间（首签时间 - 末次渠道更新日期），单位:天
    3. 在 首签时间 列后新增一列
    """
    OUTPUT_DIR.mkdir(exist_ok=True)

    df = pd.read_excel(SALES_FILE, header=SALES_HEADER_ROW)
    print(f"[Step1] 销售明细原始数据: {len(df)} 行")

    col_channel = df.columns[SALES_COL_CHANNEL]
    df_agent = df[df[col_channel].str.contains(agent_name, na=False)].copy()
    print(f"[Step1] 筛选代理 '{agent_name}': {len(df_agent)} 行")

    if len(df_agent) == 0:
        return df_agent

    col_sign_time = df.columns[SALES_COL_SIGN_TIME]
    col_channel_date = df.columns[SALES_COL_CHANNEL_DATE]
    df_agent[col_sign_time] = pd.to_datetime(df_agent[col_sign_time], errors="coerce")
    df_agent[col_channel_date] = pd.to_datetime(df_agent[col_channel_date], errors="coerce")

    time_diff = (df_agent[col_sign_time] - df_agent[col_channel_date]).dt.days
    insert_pos = SALES_COL_SIGN_TIME + 1
    df_agent.insert(insert_pos, COL_TRANSIT_TIME, time_diff)
    print(f"[Step1] 进线-成交时间范围: {time_diff.min()} ~ {time_diff.max()} 天")

    return df_agent


def step2_filter_refund_orders(agent_name: str, refund_month: str) -> pd.DataFrame:
    """
    Step 2: 从主订单宽表筛选指定代理的本月退费订单
    """
    df = pd.read_excel(ORDERS_FILE, header=ORDERS_HEADER_ROW)
    print(f"[Step2] 主订单宽表原始数据: {len(df)} 行")

    col_channel = df.columns[ORDERS_COL_CHANNEL]
    df_agent = df[df[col_channel].str.contains(agent_name, na=False)].copy()
    print(f"[Step2] 筛选代理 '{agent_name}': {len(df_agent)} 行")

    col_refund_month = df.columns[ORDERS_COL_REFUND_MONTH]
    df_refund = df_agent[df_agent[col_refund_month] == refund_month].copy()
    print(f"[Step2] 筛选退费月份 '{refund_month}': {len(df_refund)} 行")

    col_refund_type = df.columns[ORDERS_COL_REFUND_TYPE]
    valid_types = ["未开课退费", "试学期内退费"]
    df_final = df_refund[df_refund[col_refund_type].isin(valid_types)].copy()
    print(f"[Step2] 筛选退费类型（未开课/试学期内）: {len(df_final)} 行")

    return df_final


def build_channel_summary(df_sales: pd.DataFrame) -> pd.DataFrame:
    """
    渠道汇总：进线-成交时间<60 → 按末次渠道名称聚合，对首签金额求和
    返回 DataFrame(末次渠道名称, 求和:首签金额)
    """
    # 筛选 进线-成交时间 < 60
    df_filter = df_sales[df_sales[COL_TRANSIT_TIME] < 60].copy()

    # 找到"末次渠道名称"和"首签金额"两列
    channel_col = None
    amount_col = None
    for col in df_filter.columns:
        if str(col).strip() == "末次渠道名称":
            channel_col = col
        if str(col).strip() == COL_SIGN_AMOUNT:
            amount_col = col
    if channel_col is None or amount_col is None:
        print(f"[警告] 找不到渠道列或首签金额列")
        print(f"  channel_col={channel_col}, amount_col={amount_col}")
        return pd.DataFrame(columns=["末次渠道名称", "求和:首签金额"])

    summary = df_filter.groupby(channel_col)[amount_col].sum().reset_index()
    summary.columns = ["末次渠道名称", "求和:首签金额"]
    return summary


def find_sheet_by_keyword(wb, keyword: str):
    """在工作簿中按关键字模糊匹配 sheet 名"""
    for name in wb.sheetnames:
        if keyword in name:
            return wb[name], name
    return None, None


def step3_build_settlement(df_sales: pd.DataFrame, df_refunds: pd.DataFrame,
                            agent_name: str, last_month_info: tuple, cps_rate: float = 0.35) -> Path:
    """
    Step 3: 生成结算表
    复制模板 → 填入成交明细 → 填入渠道汇总 → 填入退费明细 → 填入CPS服务费表

    Args:
        cps_rate: CPS 费率，默认 0.35
    """
    ym, m_num, m_display, period = last_month_info

    # 输出文件名
    yy = ym.split("-")[0][-2:]  # "26"
    out_name = f"香港大v {agent_name}{ym.split('-')[0]}年{m_num}月渠道结算单.xlsx"
    out_path = OUTPUT_DIR / out_name

    # 复制模板
    shutil.copy(TEMPLATE_FILE, out_path)
    print(f"[Step3] 已复制模板 → {out_path.name}")

    wb = load_workbook(out_path)
    print(f"[Step3] 模板 sheets: {wb.sheetnames}")

    # ─── 3.1 重命名并填充成交明细 sheet ───
    # 思路：删除模板中的旧成交 sheet，新建一个，然后写入完整的销售明细
    sales_sheet, sales_old_name = find_sheet_by_keyword(wb, "成交时间")
    new_sales_name = f"思维-{agent_name}成交时间{yy}年{m_num}月"
    sales_sheet_idx = None
    if sales_sheet:
        sales_sheet_idx = wb.sheetnames.index(sales_old_name)
        del wb[sales_old_name]
    sales_sheet = wb.create_sheet(new_sales_name, index=sales_sheet_idx if sales_sheet_idx is not None else 1)

    # 写入表头（第8行，模拟模板结构）
    for col_idx, col_name in enumerate(df_sales.columns, start=1):
        sales_sheet.cell(8, col_idx).value = str(col_name)

    # 写入数据（从第9行）
    data_start_row = 9
    for row_idx, (_, row_data) in enumerate(df_sales.iterrows(), start=data_start_row):
        for col_idx, val in enumerate(row_data.values, start=1):
            if pd.isna(val):
                sales_sheet.cell(row_idx, col_idx).value = None
            elif isinstance(val, pd.Timestamp):
                sales_sheet.cell(row_idx, col_idx).value = val.to_pydatetime()
            else:
                sales_sheet.cell(row_idx, col_idx).value = val

    # 标注"进线-成交时间"列为黄色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    transit_col_idx = None
    for col_idx, col_name in enumerate(df_sales.columns, start=1):
        if col_name == COL_TRANSIT_TIME:
            transit_col_idx = col_idx
            break
    if transit_col_idx:
        sales_sheet.cell(8, transit_col_idx).fill = yellow_fill
        for r in range(data_start_row, data_start_row + len(df_sales)):
            sales_sheet.cell(r, transit_col_idx).fill = yellow_fill
    print(f"[Step3.1] 成交明细已写入 sheet [{new_sales_name}]，共 {len(df_sales)} 行")

    # ─── 3.2 填充渠道汇总 sheet ───
    summary_sheet, summary_name = find_sheet_by_keyword(wb, "渠道汇总")
    if summary_sheet:
        summary_df = build_channel_summary(df_sales)
        print(f"[Step3.2] 渠道汇总：{len(summary_df)} 个渠道，进线-成交时间<60")

        # 清空原有数据
        max_clear_row = summary_sheet.max_row
        for r in range(1, max_clear_row + 1):
            for c in range(1, summary_sheet.max_column + 1):
                summary_sheet.cell(r, c).value = None

        # 写入透视表头
        summary_sheet.cell(1, 1).value = "末次渠道名称"
        summary_sheet.cell(1, 2).value = "求和项:首签金额"
        summary_sheet.cell(2, 1).value = "(进线-成交时间筛选: <60)"
        summary_sheet.cell(3, 1).value = "末次渠道名称"
        summary_sheet.cell(3, 2).value = "求和项:首签金额"

        # 写入数据
        start_row = 4
        for i, row in summary_df.iterrows():
            summary_sheet.cell(start_row + i, 1).value = row["末次渠道名称"]
            summary_sheet.cell(start_row + i, 2).value = row["求和:首签金额"]

        # 总计行
        total_row = start_row + len(summary_df)
        summary_sheet.cell(total_row, 1).value = "总计"
        summary_sheet.cell(total_row, 2).value = summary_df["求和:首签金额"].sum()

    # ─── 3.3 填充退费明细 sheet ───
    # 重命名/或新建退费明细 sheet
    refund_sheet_name = f"思维-{agent_name}退费时间{yy}年{m_num}月"
    if refund_sheet_name in wb.sheetnames:
        refund_sheet = wb[refund_sheet_name]
    else:
        # 模板里没有退费 sheet，新建一个
        refund_sheet = wb.create_sheet(refund_sheet_name)

    if len(df_refunds) == 0:
        refund_sheet.cell(1, 1).value = "本月暂无退费"
        print(f"[Step3.3] 退费明细 sheet [{refund_sheet_name}]：本月暂无退费")
    else:
        # 写入表头
        for col_idx, col_name in enumerate(df_refunds.columns, start=1):
            refund_sheet.cell(1, col_idx).value = str(col_name)
        # 写入数据
        for row_idx, (_, row_data) in enumerate(df_refunds.iterrows(), start=2):
            for col_idx, val in enumerate(row_data.values, start=1):
                if pd.isna(val):
                    refund_sheet.cell(row_idx, col_idx).value = None
                elif isinstance(val, pd.Timestamp):
                    refund_sheet.cell(row_idx, col_idx).value = val.to_pydatetime()
                else:
                    refund_sheet.cell(row_idx, col_idx).value = val
        print(f"[Step3.3] 退费明细 sheet [{refund_sheet_name}]：{len(df_refunds)} 行")

    # ─── 3.4 填充 CPS 服务费表 ───
    cps_sheet, cps_old_name = find_sheet_by_keyword(wb, "CPS")
    if cps_sheet:
        # 计算上个月日期范围
        ym_year = ym.split("-")[0]
        from calendar import monthrange
        last_day = monthrange(int(ym_year), m_num)[1]
        period_in_cps = f"{ym_year}.{m_num}.1-{m_num}.{last_day}"

        # 修改标题行 B2: "【KOL-Amy】香港地區2026年4月份結算費用清單"
        cps_sheet.cell(2, 2).value = f"【KOL-{agent_name}】香港地區{ym_year}年{m_num}月份結算費用清單"

        # 准备退费金额映射表：渠道名称 → 退费金额（按渠道求和）
        refund_map = {}
        if len(df_refunds) > 0:
            refund_channel_col = df_refunds.columns[ORDERS_COL_CHANNEL]
            refund_amount_col = df_refunds.columns[ORDERS_COL_REFUND_AMOUNT]
            for ch, amt in df_refunds.groupby(refund_channel_col)[refund_amount_col].sum().items():
                refund_map[ch] = amt

        # 准备渠道汇总数据
        summary_df = build_channel_summary(df_sales)

        # 清空原有 CPS 数据（行7~31）
        for r in range(7, 32):
            for c in range(2, 13):
                cps_sheet.cell(r, c).value = None

        # 写入新数据
        data_start = 7
        used_channels = set()
        for i, row in summary_df.iterrows():
            r = data_start + i
            channel = row["末次渠道名称"]
            amount = row["求和:首签金额"]
            refund_amt = refund_map.get(channel, 0)
            used_channels.add(channel)

            cps_sheet.cell(r, 2).value = period_in_cps  # 资料期间
            cps_sheet.cell(r, 3).value = "思維"  # 主投科目
            cps_sheet.cell(r, 4).value = channel  # 管道名稱
            cps_sheet.cell(r, 6).value = amount  # 總訂單金額
            cps_sheet.cell(r, 7).value = refund_amt  # 退款金額
            cps_sheet.cell(r, 8).value = f"=F{r}-G{r}"  # 有效订单金额
            cps_sheet.cell(r, 9).value = cps_rate  # CPS資金率
            cps_sheet.cell(r, 10).value = f"=H{r}*I{r}"  # CPS服务费(元)
            cps_sheet.cell(r, 11).value = 0.87581  # 汇率（占位，3.5 步骤会更新）
            cps_sheet.cell(r, 12).value = f"=J{r}/K{r}"  # CPS服务费(港幣)

        # 处理退费表中存在但渠道汇总里没有的渠道（新建行）
        unmatched = [ch for ch in refund_map.keys() if ch not in used_channels]
        for j, ch in enumerate(unmatched):
            r = data_start + len(summary_df) + j
            cps_sheet.cell(r, 2).value = period_in_cps
            cps_sheet.cell(r, 3).value = "思維"
            cps_sheet.cell(r, 4).value = ch
            cps_sheet.cell(r, 6).value = 0  # 没有成交
            cps_sheet.cell(r, 7).value = refund_map[ch]
            cps_sheet.cell(r, 8).value = f"=F{r}-G{r}"
            cps_sheet.cell(r, 9).value = 0.35
            cps_sheet.cell(r, 10).value = f"=H{r}*I{r}"
            cps_sheet.cell(r, 11).value = 0.87581
            cps_sheet.cell(r, 12).value = f"=J{r}/K{r}"

        total_data_rows = len(summary_df) + len(unmatched)
        print(f"[Step3.4] CPS 服务费表已填充 {total_data_rows} 行（成交 {len(summary_df)} + 新增退费 {len(unmatched)}）")
        print(f"  期间: {period_in_cps}")
    else:
        print(f"[警告] 未找到 CPS sheet")

    # ─── 3.5 重命名汇率 sheet（截图 + 汇率回填在 step3_5 中处理）───
    rate_sheet, rate_old_name = find_sheet_by_keyword(wb, "汇率")
    if rate_sheet:
        new_rate_name = f"汇率{m_num}月"
        if rate_old_name != new_rate_name:
            rate_sheet.title = new_rate_name
        print(f"[Step3.5] 汇率 sheet 重命名为 [{new_rate_name}]")

    wb.save(out_path)
    print(f"[Step3] 已保存(初版): {out_path}")
    return out_path


def step3_5_fetch_exchange_rate(out_path: Path, last_month_info: tuple) -> None:
    """
    Step 3.5: 从国家外汇局抓取上月最后一个工作日的汇率
    1. 计算上月最后一个工作日（遇周末向前回溯）
    2. Playwright 打开 https://www.safe.gov.cn/safe/rmbhlzjj/index.html
    3. 设置日期 → 查询 → 截取汇率表区域
    4. 提取「港元」列的值（100港元=X人民币 → 折算成 1人民币=Y港元，Y=港元值/100）
    5. 截图插入到「汇率{m}月」sheet，并更新 CPS K 列汇率
    """
    from playwright.sync_api import sync_playwright

    ym, m_num, m_display, period = last_month_info
    ym_year = int(ym.split("-")[0])
    target_date = get_last_working_day(ym_year, m_num)
    target_date_str = target_date.strftime("%Y-%m-%d")
    print(f"[Step3.5] 目标日期（上月最后工作日）: {target_date_str} (周{['一','二','三','四','五','六','日'][target_date.weekday()]})")

    screenshot_path = OUTPUT_DIR / f"rate_{ym}.png"
    hkd_per_100_rmb = None  # 100人民币兑港元 (页面上是这个口径吗？需要看页面)
    rate_for_cps = None  # 1人民币 = X 港元

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(viewport={"width": 1400, "height": 900})
        page = context.new_page()
        try:
            print(f"[Step3.5] 打开外汇局页面...")
            page.goto("https://www.safe.gov.cn/safe/rmbhlzjj/index.html", timeout=60000)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(3000)

            # 找到查询表单所在的 iframe
            iframe = None
            for f in page.frames:
                if "RMBQuery.do" in f.url:
                    iframe = f
                    break
            if not iframe:
                print("[Step3.5] 未找到查询 iframe，终止")
                return

            print(f"[Step3.5] 已定位到 iframe: {iframe.url}")

            # 在 iframe 中设置日期范围
            try:
                start_input = iframe.locator("input#startDateId")
                end_input = iframe.locator("input#endDateId")
                start_input.fill(target_date_str)
                end_input.fill(target_date_str)
                print(f"[Step3.5] 已设置日期范围: {target_date_str}")
                iframe.wait_for_timeout(500)
            except Exception as e:
                print(f"[Step3.5] 日期输入框定位异常: {e}")

            # 点击查询按钮（第一个 input[type='button']）
            try:
                btn = iframe.locator("input[type='button']").first
                btn.click()
                print("[Step3.5] 已点击查询按钮")
                iframe.wait_for_timeout(3000)
            except Exception as e:
                print(f"[Step3.5] 查询按钮点击异常: {e}")

            # 等待表格出现
            iframe.wait_for_selector("table", timeout=15000)
            iframe.wait_for_timeout(1500)

            # 截图：取 iframe 中的表格区域
            tables = iframe.locator("table").all()
            print(f"[Step3.5] iframe 中找到 {len(tables)} 个表格")
            # 找到列数最多的表格（汇率数据表通常有20+列）
            target_table = None
            max_cols = 0
            for t in tables:
                try:
                    rows = t.locator("tr").count()
                    if rows > 1:
                        cols = t.locator("tr").first.locator("td, th").count()
                        if cols > max_cols:
                            max_cols = cols
                            target_table = t
                except Exception:
                    continue
            print(f"[Step3.5] 选中数据表格：{target_table.locator('tr').count() if target_table else 0} 行 × {max_cols} 列")

            if target_table:
                try:
                    target_table.screenshot(path=str(screenshot_path))
                    print(f"[Step3.5] 已保存截图: {screenshot_path.name}")
                except Exception as e:
                    print(f"[Step3.5] 表格截图失败，尝试 iframe 截图: {e}")
                    iframe.locator("body").screenshot(path=str(screenshot_path))

            # 提取「港元」列数值
            try:
                if target_table:
                    headers = target_table.locator("tr").first.locator("td, th").all_inner_texts()
                    print(f"[Step3.5] 表头: {headers}")
                    hk_idx = None
                    for i, h in enumerate(headers):
                        if "港元" in h or "港币" in h:
                            hk_idx = i
                            break
                    if hk_idx is None:
                        print("[Step3.5] 未识别到港元列，请检查截图后手填 K 列")
                    else:
                        rows = target_table.locator("tr").all()
                        for row in rows[1:]:
                            cells = row.locator("td").all_inner_texts()
                            if len(cells) > hk_idx and cells[hk_idx].strip():
                                try:
                                    hkd_per_100_rmb = float(cells[hk_idx].replace(",", "").strip())
                                    print(f"[Step3.5] 港元值（100人民币兑港元）: {hkd_per_100_rmb}")
                                    break
                                except ValueError:
                                    continue
            except Exception as e:
                print(f"[Step3.5] 表格解析失败: {e}")

        finally:
            page.wait_for_timeout(1000)
            browser.close()

    if hkd_per_100_rmb:
        # 用户口径：1人民币 = (港元值/100) 港元
        rate_for_cps = round(hkd_per_100_rmb / 100, 5)
        print(f"[Step3.5] 折算后 1人民币 = {rate_for_cps} 港元")

    # 把截图和汇率写回 Excel
    wb = load_workbook(out_path)
    rate_sheet, rate_name = find_sheet_by_keyword(wb, "汇率")
    if rate_sheet and screenshot_path.exists():
        try:
            img = XLImage(str(screenshot_path))
            img.anchor = "A1"
            rate_sheet.add_image(img)
            print(f"[Step3.5] 截图已插入 sheet [{rate_name}]")
        except Exception as e:
            print(f"[Step3.5] 截图插入失败: {e}")

    if rate_for_cps:
        cps_sheet, _ = find_sheet_by_keyword(wb, "CPS")
        if cps_sheet:
            for r in range(7, 32):
                if cps_sheet.cell(r, 4).value:  # 有渠道名才写
                    cps_sheet.cell(r, 11).value = rate_for_cps
            print(f"[Step3.5] CPS K列汇率已更新为 {rate_for_cps}")

    wb.save(out_path)
    print(f"[Step3.5] 已保存最终版: {out_path}")


def main():
    # 从命令行参数读取代理名称和 CPS 率
    # 用法：python process_settlement.py [代理名称] [CPS率]
    # 示例：python process_settlement.py 船长 0.4
    agent_name = sys.argv[1] if len(sys.argv) > 1 else "Amy"
    cps_rate = float(sys.argv[2]) if len(sys.argv) > 2 else 0.35

    last_month_info = get_last_month_info()
    ym, m_num, m_display, period = last_month_info
    print(f"\n[配置] 代理: {agent_name}, CPS率: {cps_rate}, 结算月份: {ym} ({m_display})\n")

    df_sales = step1_extract_agent_sales(agent_name)
    df_refunds = step2_filter_refund_orders(agent_name, ym)

    if len(df_sales) == 0:
        print("[错误] 销售明细为空，终止")
        return

    out_path = step3_build_settlement(df_sales, df_refunds, agent_name, last_month_info, cps_rate)
    step3_5_fetch_exchange_rate(out_path, last_month_info)
    print(f"\n[完成] 输出: {out_path}")


if __name__ == "__main__":
    main()
